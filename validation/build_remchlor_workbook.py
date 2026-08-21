"""Build REMChlor_validation.xlsx — a workbook to validate the rate estimators against REMChlor.

REMChlor cannot be driven headlessly from here (its GUI needs a desktop session and the Fortran
engine takes an undocumented NAMELIST input the GUI writes at run time). So this builds a
self-contained workbook the user runs alongside REMChlor:

  * Method 2 (Domenico-normalized flowpath lambda) is the apples-to-apples test: REMChlor's plume
    IS a Domenico-family streamtube model with a first-order decay rate k, so Method 2 should
    recover k. The workbook does this fully in Excel (ERF normalization + log-linear slope +
    inversion) and ALSO shows the raw, un-normalized slope, which over-estimates because it
    folds transverse dilution into the decay -- demonstrating exactly what normalization fixes.

  * Methods 1 and 3 measure DIFFERENT quantities (temporal/source decline and whole-plume mass
    loss) and are not expected to equal REMChlor's plume k -- which reinforces the three-estimand
    rule. They are driven by a second (decaying-source) scenario and the Python harness.

Run:  python validation/build_remchlor_workbook.py
"""

from __future__ import annotations

import math
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "outputs", "validation")

# ---- the prescribed Scenario A (clean Method-2 test) -------------------------------------
SCEN_A = dict(
    C0=1.0, M0=1.0e6, gamma=1.0, Y=10.0, Z=3.0,
    vd=30.0, porosity=0.30, R=1.0, sigmav=0.05, vmin=0.7, vmax=1.3, ntubes=100,
    alphay=1.0, alphaz=0.1, lambda_true=0.5, snapshot_year=20,
    x_min=25, x_max=400, x_step=25,
)
SCEN_A["v_pore"] = SCEN_A["vd"] / SCEN_A["porosity"]      # 100 m/yr
SCEN_A["vc"] = SCEN_A["v_pore"] / SCEN_A["R"]             # 100 m/yr
SCEN_A["alpha_x_infer"] = 1.0                            # small (sigmav small); correction term tiny

HDR = Font(bold=True, size=12, color="FFFFFF")
HDRFILL = PatternFill("solid", fgColor="1F4E79")
SUB = Font(bold=True, size=11, color="1F4E79")
INFILL = PatternFill("solid", fgColor="FFF2CC")          # yellow = user enters / pastes
OUTFILL = PatternFill("solid", fgColor="E2EFDA")         # green = computed result
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="BBBBBB")] * 4)


def _title(ws, text):
    ws["A1"] = text
    ws["A1"].font = HDR; ws["A1"].fill = HDRFILL
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 22


def build():
    wb = Workbook()

    # ============================================================ 1. README
    ws = wb.active; ws.title = "1_README"; _title(ws, "REMChlor validation of the bulk-attenuation rate estimators")
    lines = [
        ("", ""),
        ("Why this workbook", SUB),
        ("REMChlor runs only through its Windows GUI (the Fortran engine reads an undocumented "
         "namelist the GUI writes at run time), so it cannot be driven automatically. Instead you "
         "run REMChlor yourself with the prescribed scenario, paste its output here, and the "
         "workbook computes our estimate and compares it to the rate REMChlor used.", WRAP),
        ("", ""),
        ("The key idea", SUB),
        ("REMChlor's plume is a Domenico-family streamtube model with a first-order plume decay "
         "rate k. Method 2 (Domenico-normalized) inverts exactly that model, so it should RECOVER "
         "k. This is the rigorous test (sheet 4). Methods 1 and 3 measure different quantities "
         "(temporal/source decline; whole-plume mass loss) and are NOT expected to equal k -- that "
         "is the point of keeping the three estimates separate and never averaging them.", WRAP),
        ("", ""),
        ("What you do", SUB),
        ("1. Open REMChlor. Build a project with the inputs on sheet '2_ScenarioA_Setup' "
         "(a clean, steady-source, single-component run with a known uniform plume decay rate).", WRAP),
        ("2. Run it. Open the output file REMChlor.csv (comma-delimited: time, x, y, z, C1..C4, "
         "Ctotal).", WRAP),
        ("3. Keep only the centerline rows (y = 0, z = 0) at the snapshot time, and paste their "
         "x and concentration into sheet '3_PasteA_Centerline'.", WRAP),
        ("4. Read the verdict on sheet '4_Method2_vs_REMChlor': the normalized lambda should match "
         "REMChlor's k; the raw (un-normalized) lambda will be biased high.", WRAP),
        ("5. (Optional) For Methods 1 and 3, run Scenario B (sheet 6), paste the time series on "
         "sheet 7, and either read the in-cell Method-1 rate or run the Python harness "
         "'biodeg_rates/validate_remchlor.py' on this saved workbook for the exact three-method run.", WRAP),
        ("", ""),
        ("Colour key", SUB),
        ("Yellow cells = you enter or paste.   Green cells = computed result.", WRAP),
    ]
    r = 2
    for text, style in lines:
        ws.cell(r, 1, text)
        if style is SUB:
            ws.cell(r, 1).font = SUB
        elif style is WRAP:
            ws.cell(r, 1).alignment = WRAP
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            ws.row_dimensions[r].height = 42
        r += 1
    ws.column_dimensions["A"].width = 22
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 12

    # ============================================================ 2. Scenario A setup
    ws = wb.create_sheet("2_ScenarioA_Setup"); _title(ws, "Scenario A — REMChlor inputs (Method 2 test): steady source, known plume k")
    ws["A3"] = ("Enter these in REMChlor. The source is made effectively constant (huge mass) so the "
                "plume reaches steady state; every plume decay-rate cell is set to the same known k.")
    ws["A3"].alignment = WRAP; ws.merge_cells("A3:H3"); ws.row_dimensions[3].height = 30
    rows = [
        ("REMChlor field", "Value", "Units", "Notes"),
        ("Initial source concentration C0", SCEN_A["C0"], "mg/L", "any value; the slope, not the level, carries k"),
        ("Source mass M0", SCEN_A["M0"], "kg", "deliberately huge -> source ~ constant -> steady plume"),
        ("Source decay exponent Gamma", SCEN_A["gamma"], "-", "irrelevant while M0 is huge"),
        ("Source width Y", SCEN_A["Y"], "m", "transverse source dimension (used by the normalization)"),
        ("Source height Z", SCEN_A["Z"], "m", "vertical source dimension"),
        ("Darcy velocity vd", SCEN_A["vd"], "m/yr", ""),
        ("Porosity", SCEN_A["porosity"], "-", "pore velocity v = vd / porosity = %.0f m/yr" % SCEN_A["v_pore"]),
        ("Retardation R", SCEN_A["R"], "-", "keep 1 for a clean test; vc = v/R = %.0f m/yr" % SCEN_A["vc"]),
        ("Sigma v", SCEN_A["sigmav"], "-", "small -> minimal longitudinal dispersion (clean exponential)"),
        ("v min / v max", "%.1f / %.1f" % (SCEN_A["vmin"], SCEN_A["vmax"]), "-", "streamtube velocity bounds"),
        ("Number of streamtubes", SCEN_A["ntubes"], "-", ""),
        ("Transverse dispersivity alpha_y", SCEN_A["alphay"], "m", "PRESENT on purpose: it dilutes the centerline, which the normalization must remove"),
        ("Vertical dispersivity alpha_z", SCEN_A["alphaz"], "m", ""),
        ("Yields (3 to 2, 4 to 3, etc.)", 0, "-", "0 -> single component, no daughters"),
        ("PLUME DECAY RATE k (all 9 cells, Component 1)", SCEN_A["lambda_true"], "1/yr", "THE KNOWN TRUTH. Set every zone x period cell to this."),
        ("Zone distances X1, X2", "1000 / 2000", "m", "beyond the plume, so a single uniform zone applies"),
        ("Time periods t1, t2", "1000 / 1000", "yr", "huge, so a single uniform period applies"),
        ("Source remediation", "none", "-", "no remediation in this scenario"),
        ("Output snapshot time", SCEN_A["snapshot_year"], "yr", "plume is fully stabilized (front at %d m)" % int(SCEN_A["v_pore"] * SCEN_A["snapshot_year"])),
        ("Output x range", "%d to %d step %d" % (SCEN_A["x_min"], SCEN_A["x_max"], SCEN_A["x_step"]), "m", "centerline transect to paste"),
        ("Output y, z", "0, 0", "m", "centerline, base of source"),
    ]
    for i, row in enumerate(rows):
        rr = 5 + i
        for j, val in enumerate(row):
            c = ws.cell(rr, 1 + j, val)
            c.border = THIN
            if i == 0:
                c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D6E4F0")
            if j == 1 and i > 0:
                c.fill = INFILL
            if j == 3:
                c.alignment = WRAP
    ws.column_dimensions["A"].width = 38; ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 8; ws.column_dimensions["D"].width = 52

    # ============================================================ 3. Paste centerline
    ws = wb.create_sheet("3_PasteA_Centerline"); _title(ws, "Paste REMChlor centerline output (Scenario A snapshot, y=0, z=0)")
    ws["A3"] = ("Paste the centerline rows from REMChlor.csv at the snapshot time: distance x (m) in "
                "column A and the (total) concentration in column B. Start at row 6. Up to ~50 rows.")
    ws["A3"].alignment = WRAP; ws.merge_cells("A3:F3"); ws.row_dimensions[3].height = 30
    heads = ["x (m)  [paste]", "C (mg/L)  [paste]", "Phi_y (computed)", "ln C* normalized", "ln C raw"]
    for j, h in enumerate(heads):
        c = ws.cell(5, 1 + j, h); c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D6E4F0"); c.border = THIN
    first, last = 6, 55
    M = "'4_Method2_vs_REMChlor'"
    for rr in range(first, last + 1):
        ws.cell(rr, 1).fill = INFILL; ws.cell(rr, 2).fill = INFILL
        # Phi_y at y=0: 0.5*(erf((Y/2)/(2 sqrt(ay x))) - erf((-Y/2)/(2 sqrt(ay x)))) = erf((Y/2)/(2 sqrt(ay x)))
        ws.cell(rr, 3, f'=IF(ISNUMBER(A{rr}),ERF(({M}!$B$5/2)/(2*SQRT({M}!$B$6*A{rr}))),"")')
        ws.cell(rr, 4, f'=IF(AND(ISNUMBER(A{rr}),ISNUMBER(B{rr})),LN(B{rr}/C{rr}),"")')
        ws.cell(rr, 5, f'=IF(ISNUMBER(B{rr}),LN(B{rr}),"")')
    for col, w in (("A", 14), ("B", 16), ("C", 16), ("D", 16), ("E", 14)):
        ws.column_dimensions[col].width = w

    # ============================================================ 4. Method 2 result
    ws = wb.create_sheet("4_Method2_vs_REMChlor"); _title(ws, "Method 2 (Domenico-normalized) vs REMChlor plume decay rate k")
    P = "'3_PasteA_Centerline'"
    inputs = [
        ("Known transport parameters (from Scenario A)", "", ""),
        ("Pore velocity v", SCEN_A["v_pore"], "m/yr"),
        ("Source width Y", SCEN_A["Y"], "m"),
        ("Transverse dispersivity alpha_y", SCEN_A["alphay"], "m"),
        ("Retardation R", SCEN_A["R"], "-"),
        ("Contaminant velocity vc = v/R", f"=B3/B6", "m/yr"),
        ("Longitudinal dispersivity alpha_x (small)", SCEN_A["alpha_x_infer"], "m"),
        ("REMChlor plume decay rate k (TRUTH)", SCEN_A["lambda_true"], "1/yr"),
    ]
    # place inputs at B3..B10 (note: B5=Y, B6 will be alpha_y? need consistent refs used in sheet 3)
    # Lay out: A2 label header
    ws["A2"] = "Inputs"; ws["A2"].font = SUB
    layout = [
        (3, "Pore velocity v (m/yr)", SCEN_A["v_pore"]),
        (4, "Retardation R", SCEN_A["R"]),
        (5, "Source width Y (m)", SCEN_A["Y"]),
        (6, "Transverse dispersivity alpha_y (m)", SCEN_A["alphay"]),
        (7, "Longitudinal dispersivity alpha_x (m)", SCEN_A["alpha_x_infer"]),
        (8, "REMChlor plume decay rate k  (TRUTH, 1/yr)", SCEN_A["lambda_true"]),
    ]
    for rr, lab, val in layout:
        ws.cell(rr, 1, lab); c = ws.cell(rr, 2, val); c.fill = INFILL; c.border = THIN
    ws.cell(9, 1, "Contaminant velocity vc = v/R (m/yr)")
    ws.cell(9, 2, "=B3/B4").fill = OUTFILL

    ws["A12"] = "Results"; ws["A12"].font = SUB
    res = [
        (13, "Normalized slope m (1/m)", f"=SLOPE({P}!D6:D55,{P}!A6:A55)"),
        (14, "Raw slope m_raw (1/m)", f"=SLOPE({P}!E6:E55,{P}!A6:A55)"),
        (15, "lambda recovered, NORMALIZED (1/yr)", "=-B9*B13*(1-B7*B13)"),
        (16, "lambda from RAW slope, no dilution removal (1/yr)", "=-B9*B14*(1-B7*B14)"),
        (17, "lambda_true from REMChlor (1/yr)", "=B8"),
        (18, "error, normalized (%)", "=100*(B15-B17)/B17"),
        (19, "error, raw (%)", "=100*(B16-B17)/B17"),
    ]
    for rr, lab, f in res:
        ws.cell(rr, 1, lab); c = ws.cell(rr, 2, f); c.fill = OUTFILL; c.border = THIN
        c.number_format = "0.000"
    ws.cell(21, 1, "VERDICT").font = Font(bold=True)
    ws.cell(21, 2, '=IF(ABS(B18)<=20,"PASS: normalized Method 2 recovers REMChlor k within 20%",'
                   '"REVIEW: see inputs and pasted data")')
    ws.cell(21, 2).font = Font(bold=True)
    ws.cell(22, 1, "Interpretation").font = SUB
    ws.cell(23, 1, ("The NORMALIZED lambda removes transverse dilution and should match k. The RAW "
                    "lambda ignores dilution (the centerline drops partly because the plume spreads) "
                    "and over-estimates k -- this is the bias Method 2 is built to remove.")).alignment = WRAP
    ws.merge_cells("A23:H23"); ws.row_dimensions[23].height = 42
    ws.column_dimensions["A"].width = 44; ws.column_dimensions["B"].width = 18

    # ============================================================ 5. Scenario B setup
    ws = wb.create_sheet("5_ScenarioB_Setup"); _title(ws, "Scenario B — decaying source (for Methods 1 and 3, temporal)")
    ws["A3"] = ("Methods 1 (per-well point decay) and 3 (plume mass loss) measure how concentrations "
                "fall over TIME, driven mainly by source depletion -- a different quantity than the "
                "plume reaction rate k. Run a second REMChlor scenario with a finite, decaying source "
                "and sample several wells over time.")
    ws["A3"].alignment = WRAP; ws.merge_cells("A3:H3"); ws.row_dimensions[3].height = 46
    rowsB = [
        ("REMChlor field", "Value", "Units", "Notes"),
        ("Initial source concentration C0", 2.0, "mg/L", ""),
        ("Source mass M0", 1620, "kg", "finite -> source depletes over time"),
        ("Source decay exponent Gamma", 1.0, "-", "near-linear-in-log decline"),
        ("Darcy velocity vd", SCEN_A["vd"], "m/yr", ""),
        ("Porosity / R", "0.30 / 1", "-", ""),
        ("alpha_y / alpha_z", "1.0 / 0.1", "m", ""),
        ("Plume decay rate k (all cells)", SCEN_A["lambda_true"], "1/yr", "same k as Scenario A"),
        ("Output well distances x", "25, 50, 100, 150, 200, 300", "m", "centerline monitoring wells"),
        ("Output times", "1,3,5,8,12,16,20,25,30", "yr", "the monitoring rounds"),
        ("EXPECTATION", "Method 1 ~ source decline rate; Method 3 ~ plume mass loss; neither equals k", "", "demonstrates the three estimands differ"),
    ]
    for i, row in enumerate(rowsB):
        rr = 5 + i
        for j, val in enumerate(row):
            c = ws.cell(rr, 1 + j, val); c.border = THIN
            if i == 0:
                c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D6E4F0")
            if j == 1 and 0 < i < len(rowsB) - 1:
                c.fill = INFILL
            if j == 3:
                c.alignment = WRAP
    ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 8; ws.column_dimensions["D"].width = 40

    # ============================================================ 6. Paste time series + Method 1
    ws = wb.create_sheet("6_PasteB_Method1"); _title(ws, "Paste time series (one well) and read the Method-1 point-decay rate")
    ws["A3"] = ("Paste a single well's time series from Scenario B: time (yr) in column A, concentration "
                "(mg/L) in column B. The in-cell slope of ln C vs time is the point-decay rate. Repeat "
                "per well, or use the Python harness for all wells at once.")
    ws["A3"].alignment = WRAP; ws.merge_cells("A3:F3"); ws.row_dimensions[3].height = 42
    for j, h in enumerate(["t (yr) [paste]", "C (mg/L) [paste]", "ln C"]):
        c = ws.cell(5, 1 + j, h); c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D6E4F0"); c.border = THIN
    for rr in range(6, 36):
        ws.cell(rr, 1).fill = INFILL; ws.cell(rr, 2).fill = INFILL
        ws.cell(rr, 3, f'=IF(ISNUMBER(B{rr}),LN(B{rr}),"")')
    ws.cell(38, 1, "point-decay k (1/yr)")
    ws.cell(38, 2, "=-SLOPE(C6:C35,A6:A35)").fill = OUTFILL
    ws.cell(38, 2).number_format = "0.000"
    ws.cell(39, 1, "half-life (yr)")
    ws.cell(39, 2, '=IF(B38>0,LN(2)/B38,"n/a (rising)")').fill = OUTFILL
    ws.cell(41, 1, ("Note: this rate reflects SOURCE depletion seen at the well, not the plume reaction "
                    "rate k. That is expected -- it is a different estimand.")).alignment = WRAP
    ws.merge_cells("A41:H41"); ws.row_dimensions[41].height = 30
    for col, w in (("A", 18), ("B", 16), ("C", 12)):
        ws.column_dimensions[col].width = w

    # ============================================================ 7. Long format for Python
    ws = wb.create_sheet("7_LongFormat_Python"); _title(ws, "Long-format data + known params for the Python harness (all 3 methods)")
    ws["A3"] = ("For the exact production methods (Theil-Sen Method 1, Method 2, and the P-spline "
                "Method 3 which cannot run in Excel), paste every well-event here, fill the parameter "
                "block, save the workbook, and run:  python validation/validate_remchlor.py "
                "--xlsx <this file>")
    ws["A3"].alignment = WRAP; ws.merge_cells("A3:H3"); ws.row_dimensions[3].height = 42
    ws["A5"] = "Known parameters"; ws["A5"].font = SUB
    params = [("lambda_true_per_year", SCEN_A["lambda_true"]),
              ("vc_m_per_year", SCEN_A["vc"]),
              ("alpha_x_m", SCEN_A["alpha_x_infer"]),
              ("alpha_y_m", SCEN_A["alphay"]),
              ("source_width_Y_m", SCEN_A["Y"]),
              ("snapshot_year_for_method2", SCEN_A["snapshot_year"])]
    for i, (k, v) in enumerate(params):
        ws.cell(6 + i, 1, k); c = ws.cell(6 + i, 2, v); c.fill = INFILL; c.border = THIN
    hrow = 14
    for j, h in enumerate(["well_id", "easting", "northing", "t_years", "conc_mgL", "detect(TRUE/FALSE)"]):
        c = ws.cell(hrow, 1 + j, h); c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D6E4F0"); c.border = THIN
    for rr in range(hrow + 1, hrow + 1 + 300):
        for j in range(6):
            ws.cell(rr, 1 + j).fill = INFILL
    for j in range(6):
        ws.column_dimensions[get_column_letter(1 + j)].width = 16

    # ============================================================ 8. Results summary
    ws = wb.create_sheet("8_Results"); _title(ws, "Validation summary")
    ws["A3"] = "Method 2 (the apples-to-apples test) result:"
    ws["A4"] = "  normalized lambda (1/yr)"; ws["B4"] = "='4_Method2_vs_REMChlor'!B15"; ws["B4"].fill = OUTFILL
    ws["A5"] = "  REMChlor k (1/yr)"; ws["B5"] = "='4_Method2_vs_REMChlor'!B17"; ws["B5"].fill = OUTFILL
    ws["A6"] = "  verdict"; ws["B6"] = "='4_Method2_vs_REMChlor'!B21"; ws["B6"].fill = OUTFILL
    ws["A8"] = ("Methods 1 and 3 are validated by the Python harness on Scenario B; they are expected "
                "to differ from k because they measure source/temporal decline and whole-plume mass "
                "loss, not the plume reaction rate. The three estimates are never averaged.")
    ws["A8"].alignment = WRAP; ws.merge_cells("A8:H8"); ws.row_dimensions[8].height = 46
    ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 40

    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, "REMChlor_validation.xlsx")
    wb.save(path)
    return path


if __name__ == "__main__":
    p = build()
    print("workbook:", os.path.abspath(p))
